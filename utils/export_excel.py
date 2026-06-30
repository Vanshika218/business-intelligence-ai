from openpyxl import Workbook


def create_excel(sales, inventory, reviews):

    wb = Workbook()

    # ---------------- Sales Sheet ----------------

    ws = wb.active
    ws.title = "Sales"

    ws.append(["Metric", "Value"])
    ws.append(["Total Revenue", sales["total_revenue"]])
    ws.append(["Units Sold", sales["total_units"]])
    ws.append(["Top Product", sales["top_product"]])

    ws.append([])
    ws.append(["Revenue by Product"])

    for row in sales["revenue_by_product"].itertuples(index=False):
        ws.append(row)

    # ---------------- Inventory ----------------

    ws2 = wb.create_sheet("Inventory")

    for row in inventory["inventory_df"].itertuples(index=False):
        ws2.append(row)

    # ---------------- Reviews ----------------

    ws3 = wb.create_sheet("Reviews")

    ws3.append(["Metric", "Value"])
    ws3.append(["Total Reviews", reviews["total_reviews"]])
    ws3.append(["Positive Reviews", reviews["positive_reviews"]])
    ws3.append(["Negative Reviews", reviews["negative_reviews"]])
    ws3.append(["Positive %", reviews["positive_percentage"]])

    filename = "Business_Report.xlsx"

    wb.save(filename)

    return filename